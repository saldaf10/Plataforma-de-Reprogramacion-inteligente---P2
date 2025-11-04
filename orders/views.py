from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.contrib import messages
from catalog.models import Product
from .models import Order, OrderItem, Delivery, DeliveryEvent, DeliveryComment, DeliveryNotification, DeliveryFailureReason
from .notification_service import DeliveryNotificationService


@login_required
@transaction.atomic
def checkout(request):
    cart = request.session.get("cart", {})
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        address = request.POST.get("address")
        city = request.POST.get("city")
        postal_code = request.POST.get("postal_code")

        order = Order.objects.create(
            user=request.user,
            full_name=full_name,
            email=email,
            address=address,
            city=city,
            postal_code=postal_code,
        )

        total = 0
        for pid, item in cart.items():
            product = get_object_or_404(Product, id=pid)
            qty = int(item.get("quantity", 1))
            OrderItem.objects.create(order=order, product=product, price=product.price, quantity=qty)
            total += product.price * qty
        order.total_amount = total
        order.save()

        # Redirigir a simulación de pago
        request.session["last_order_id"] = order.id
        return redirect("payments:simulate")

    # GET: mostrar resumen previo
    items = []
    total = 0
    for pid, item in cart.items():
        product = get_object_or_404(Product, id=pid)
        qty = int(item.get("quantity", 1))
        line_total = product.price * qty
        total += line_total
        items.append({"product": product, "quantity": qty, "line_total": line_total})
    return render(request, "orders/checkout.html", {"items": items, "total": total})


@login_required
def order_summary(request, order_id: int):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, "orders/summary.html", {"order": order})


@login_required
def my_orders(request):
    user = request.user
    if user.is_superuser:
        return redirect("orders:panel")
    elif hasattr(user, "profile") and user.profile.role == "repartidor":
        # Mostrar entregas asignadas al rider
        deliveries = Delivery.objects.select_related("order").filter(rider=user).order_by("-created_at")
        return render(request, "orders/rider_orders.html", {"deliveries": deliveries})
    elif hasattr(user, "profile") and user.profile.role == "manager":
        return redirect("orders:panel")
    else:
        orders = Order.objects.filter(user=user).order_by("-created_at")
        # Próximo pedido no entregado
        next_delivery = (
            Delivery.objects.select_related("order")
            .filter(order__user=user)
            .exclude(status="entregada")
            .order_by("order__created_at")
            .first()
        )
        return render(request, "orders/my_orders.html", {"orders": orders, "next_delivery": next_delivery})


@login_required
def delivery_detail(request, order_id: int):
    order = get_object_or_404(Order, id=order_id)
    delivery, _ = Delivery.objects.get_or_create(order=order)
    # Cliente solo su orden; superuser/manager cualquier orden; rider si está asignado
    # Determine role
    user_role = None
    if request.user.is_superuser:
        user_role = "manager"
    elif hasattr(request.user, "profile"):
        user_role = request.user.profile.role

    if user_role == "manager":
        pass
    elif user_role == "repartidor":
        pass
    else:
        # Cliente u otro: solo su orden
        if order.user_id != request.user.id:
            return redirect("orders:my_orders")

    if request.method == "POST":
        action = request.POST.get("action")
        
        # Verificar si el pedido está en estado final
        # EXCEPCIÓN: Permitir reprogramar pedidos fallidos (clientes)
        if delivery.is_final_state:
            # Si es cliente intentando reprogramar un pedido fallido, permitir
            if action == "reschedule" and delivery.status == "fallida" and (user_role == "cliente" or not user_role):
                pass  # Permitir continuar
            else:
                messages.error(request, "Este pedido no puede ser modificado porque está en estado final (entregado o fallido).")
                return redirect("orders:delivery_detail", order_id=order.id)
        
        # Actualizaciones dependiendo del rol
        if user_role == "manager":
            if action == "assign":
                before = delivery.status
                old_rider = delivery.rider
                old_date = delivery.scheduled_date
                old_window = delivery.scheduled_window
                
                rider_id = request.POST.get("rider_id")
                new_scheduled_date = request.POST.get("scheduled_date")
                new_scheduled_window = request.POST.get("scheduled_window")
                
                # Capturar cambios antes de guardar
                date_changed = new_scheduled_date and new_scheduled_date != str(old_date) if old_date else bool(new_scheduled_date)
                window_changed = new_scheduled_window != old_window
                rider_changed = rider_id and int(rider_id) != old_rider.id if old_rider else bool(rider_id)
                
                if new_scheduled_date:
                    delivery.scheduled_date = new_scheduled_date
                if new_scheduled_window:
                    delivery.scheduled_window = new_scheduled_window
                
                if rider_id:
                    delivery.rider_id = int(rider_id)
                    delivery.status = "asignada"
                
                delivery.save()
                DeliveryEvent.objects.create(
                    delivery=delivery,
                    user=request.user,
                    status_before=before,
                    status_after=delivery.status,
                    notes=f"Asignado rider {delivery.rider.username if delivery.rider else ''}"
                )
                
                # Notificar a coordinadores sobre cambios
                # Marcar para evitar notificación duplicada desde señales
                delivery._notification_sent = True
                
                if rider_changed:
                    DeliveryNotificationService.notify_coordinators_rider_assigned(
                        delivery, old_rider=old_rider, changed_by=request.user
                    )
                
                if date_changed or window_changed:
                    DeliveryNotificationService.notify_coordinators_schedule_changed(
                        delivery, old_date=old_date, old_window=old_window, changed_by=request.user
                    )
            elif action == "manager_comment":
                msg = request.POST.get("message", "").strip()
                photo = request.FILES.get("photo")
                if msg or photo:
                    DeliveryComment.objects.create(
                        delivery=delivery,
                        user=request.user,
                        role="manager",
                        message=msg,
                        photo=photo,
                    )
        elif user_role == "repartidor":
            if action in {"en_ruta", "fallida", "entregada", "reprogramada"}:
                before = delivery.status
                
                # Validar razón de fallo si se marca como fallida
                if action == "fallida":
                    failure_reason_code = request.POST.get("failure_reason", "").strip()
                    if not failure_reason_code:
                        messages.error(request, "Debe seleccionar una razón del fallo para marcar la entrega como fallida.")
                        return redirect("orders:delivery_detail", order_id=order.id)
                
                delivery.status = action
                delivery.notes = request.POST.get("notes", "")
                photo_file = request.FILES.get("photo")
                if photo_file:
                    delivery.photo = photo_file
                delivery.save()
                
                DeliveryEvent.objects.create(
                    delivery=delivery,
                    user=request.user,
                    status_before=before,
                    status_after=delivery.status,
                    notes=delivery.notes,
                    photo=delivery.photo if photo_file else None,
                )
                
                # Si el pedido se marca como fallido, guardar la razón
                if action == "fallida":
                    from .models import DeliveryFailureReason
                    failure_reason_code = request.POST.get("failure_reason", "").strip()
                    failure_details = request.POST.get("failure_details", "").strip()
                    
                    # Calcular el número de intento (contar cuántas veces ha fallado antes)
                    attempt_number = DeliveryFailureReason.objects.filter(delivery=delivery).count() + 1
                    
                    # Guardar la razón de fallo
                    DeliveryFailureReason.objects.create(
                        delivery=delivery,
                        reason=failure_reason_code,
                        details=failure_details,
                        reported_by=request.user,
                        attempt_number=attempt_number
                    )
                    
                    # Actualizar el campo legacy para compatibilidad
                    delivery.failure_reason = failure_reason_code
                    delivery.save()
                    
                    # Construir mensaje para notificaciones
                    failure_reason_display = dict(DeliveryFailureReason.FAILURE_REASONS).get(failure_reason_code, failure_reason_code)
                    failure_message = f"{failure_reason_display}"
                    if failure_details:
                        failure_message += f": {failure_details}"
                    
                    DeliveryNotificationService.send_failed_notification(delivery, failure_message)
                    DeliveryNotificationService.notify_coordinators_delivery_failed(
                        delivery, failure_reason=failure_message, changed_by=request.user
                    )
                    messages.success(request, f"Entrega marcada como fallida (Intento #{attempt_number}). El cliente ha sido notificado automáticamente.")
                else:
                    # Marcar para evitar notificación duplicada desde señales
                    delivery._notification_sent = True
                    
                    # Notificar a coordinadores sobre cambio de estado
                    if before != delivery.status:
                        DeliveryNotificationService.notify_coordinators_status_changed(
                            delivery, old_status=before, changed_by=request.user
                        )
                    
                    messages.success(request, f"Estado actualizado a '{delivery.get_status_display()}'.")
            # Manejar notificaciones del repartidor
            elif action == "send_notification":
                # Validar que el pedido esté en un estado válido para notificaciones
                if not delivery.is_modifiable or delivery.status not in ['en_ruta', 'reprogramada']:
                    messages.error(request, "Las notificaciones solo se pueden enviar para pedidos activos en ruta o reprogramados.")
                else:
                    notification_type = request.POST.get("notification_type")
                    estimated_minutes = request.POST.get("estimated_minutes", "")
                    
                    if notification_type == "approaching":
                        minutes = int(estimated_minutes) if estimated_minutes.isdigit() else 30
                        DeliveryNotificationService.send_approaching_notification(delivery, minutes)
                        messages.success(request, f"Notificación enviada: El cliente será notificado que llegas en aproximadamente {minutes} minutos.")
                    elif notification_type == "leaving":
                        DeliveryNotificationService.send_leaving_notification(delivery)
                        messages.success(request, "Notificación enviada: El cliente fue notificado que ya saliste con su pedido.")
                    elif notification_type == "arriving_soon":
                        DeliveryNotificationService.send_arriving_soon_notification(delivery)
                        messages.success(request, "Notificación enviada: El cliente fue notificado que llegarás en 3 minutos.")
                    elif notification_type == "arrived":
                        DeliveryNotificationService.send_arrived_notification(delivery)
                        messages.success(request, "Notificación enviada: El cliente fue notificado que ya llegaste.")
                    elif notification_type == "delivered":
                        DeliveryNotificationService.send_delivered_notification(delivery)
                        messages.success(request, "Notificación enviada: El cliente fue notificado que su pedido fue entregado.")
        elif user_role == "cliente" or not user_role:
            # El cliente puede reprogramar su entrega si está en estado válido
            if action == "reschedule":
                # Solo permitir reprogramar si el estado lo permite (incluyendo fallidas)
                if delivery.status in ['pendiente', 'asignada', 'reprogramada', 'fallida']:
                    scheduled_date = request.POST.get("scheduled_date")
                    scheduled_window = request.POST.get("scheduled_window")
                    
                    if scheduled_date:
                        from datetime import datetime
                        try:
                            # Validar fecha
                            date_obj = datetime.strptime(scheduled_date, "%Y-%m-%d").date()
                            today = datetime.now().date()
                            
                            if date_obj < today:
                                messages.error(request, "No puedes seleccionar una fecha pasada.")
                            else:
                                old_date = delivery.scheduled_date
                                old_window = delivery.scheduled_window
                                
                                delivery.scheduled_date = date_obj
                                delivery.scheduled_window = scheduled_window if scheduled_window else None
                                delivery.status = "reprogramada"
                                delivery.save()
                                
                                # Crear evento
                                DeliveryEvent.objects.create(
                                    delivery=delivery,
                                    user=request.user,
                                    status_before=delivery.status,
                                    status_after="reprogramada",
                                    notes=f"Cliente reprogramó de {old_date} {old_window or ''} a {date_obj} {scheduled_window or ''}"
                                )
                                
                                # Enviar notificación al repartidor si está asignado
                                if delivery.rider:
                                    notification_message = f"📅 El cliente {request.user.username} ha reprogramado el pedido #{order.id} para el {date_obj}"
                                    if scheduled_window:
                                        notification_message += f" entre las {scheduled_window}"
                                    notification_message += ". Por favor, revisa los detalles actualizados."
                                    
                                    DeliveryNotification.objects.create(
                                        delivery=delivery,
                                        notification_type="rescheduled",
                                        recipient=delivery.rider,
                                        message=notification_message
                                    )
                                
                                # Marcar para evitar notificación duplicada desde señales
                                delivery._notification_sent = True
                                
                                # Notificar a coordinadores sobre la reprogramación
                                DeliveryNotificationService.notify_coordinators_rescheduled(
                                    delivery, old_date=old_date, old_window=old_window, changed_by=request.user
                                )
                                
                                messages.success(request, f"Tu entrega ha sido reprogramada para el {date_obj} {scheduled_window or ''}. Recibirás una confirmación pronto.")
                        except ValueError:
                            messages.error(request, "Fecha inválida. Por favor usa el formato correcto.")
                    else:
                        messages.error(request, "Debes seleccionar una fecha.")
                else:
                    messages.error(request, f"No puedes reprogramar un pedido en estado '{delivery.get_status_display()}'. Solo se pueden reprogramar pedidos pendientes, asignados, reprogramados o fallidos.")

        return redirect("orders:delivery_detail", order_id=order.id)

    # Render por rol
    from datetime import datetime
    context = {
        "order": order, 
        "delivery": delivery,
        "is_modifiable": delivery.is_modifiable,
        "is_final_state": delivery.is_final_state,
        "today": datetime.now().date()
    }
    if user_role == "manager":
        from django.contrib.auth.models import User
        riders = User.objects.filter(profile__role="repartidor")
        context["riders"] = riders
        context["events"] = delivery.events.all()
        context["comments"] = delivery.comments.all()
        return render(request, "orders/delivery_manager.html", context)
    if user_role == "repartidor":
        context["events"] = delivery.events.all()
        context["comments"] = delivery.comments.all()
        context["can_send_notifications"] = DeliveryNotificationService.can_send_notifications(delivery)
        return render(request, "orders/delivery_rider.html", context)
    # Cliente
    context["events"] = delivery.events.all()
    context["comments"] = delivery.comments.all()
    return render(request, "orders/delivery_client.html", context)


@login_required
def manager_panel(request):
    user = request.user
    if not user.is_superuser and (not hasattr(user, "profile") or user.profile.role != "manager"):
        return redirect("orders:my_orders")

    status = request.GET.get("status", "")
    q = request.GET.get("q", "")

    orders_qs = Order.objects.select_related("user").prefetch_related("items", "delivery").all()
    if status:
        orders_qs = orders_qs.filter(delivery__status=status)
    if q:
        orders_qs = orders_qs.filter(
            Q(full_name__icontains=q)
            | Q(address__icontains=q)
            | Q(city__icontains=q)
            | Q(email__icontains=q)
            | Q(id__icontains=q)
        )

    total = orders_qs.count()
    delivered = orders_qs.filter(delivery__status="entregada").count()
    failed = orders_qs.filter(delivery__status="fallida").count()
    pending = orders_qs.exclude(delivery__status="entregada").exclude(delivery__status="fallida").count()

    deliveries_by_status = Delivery.objects.values("status").annotate(c=Count("id"))

    return render(
        request,
        "orders/manager_panel.html",
        {
            "orders": orders_qs.order_by("-created_at")[:200],
            "total": total,
            "delivered": delivered,
            "failed": failed,
            "pending": pending,
            "deliveries_by_status": deliveries_by_status,
            "status": status,
            "q": q,
        },
    )


@login_required
def notifications(request):
    """Vista para mostrar las notificaciones del usuario"""
    notifications = DeliveryNotificationService.get_user_notifications(request.user, limit=20)
    
    # Marcar notificaciones como leídas si se especifica
    if request.method == "POST":
        notification_id = request.POST.get("notification_id")
        if notification_id:
            try:
                notification = DeliveryNotification.objects.get(
                    id=notification_id, 
                    recipient=request.user
                )
                DeliveryNotificationService.mark_as_read(notification)
                messages.success(request, "Notificación marcada como leída.")
            except DeliveryNotification.DoesNotExist:
                messages.error(request, "Notificación no encontrada.")
        
        return redirect("orders:notifications")
    
    return render(request, "orders/notifications.html", {
        "notifications": notifications
    })


@login_required
def download_report(request):
    """Genera y descarga reportes de entregas en Excel o PDF"""
    user = request.user
    if not user.is_superuser and (not hasattr(user, "profile") or user.profile.role != "manager"):
        return redirect("orders:my_orders")
    
    format_type = request.GET.get("format", "excel")  # excel o pdf
    status = request.GET.get("status", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    
    # Obtener entregas con filtros
    deliveries = Delivery.objects.select_related("order", "rider").all()
    
    if status:
        deliveries = deliveries.filter(status=status)
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            deliveries = deliveries.filter(scheduled_date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            deliveries = deliveries.filter(scheduled_date__lte=date_to_obj)
        except ValueError:
            pass
    
    deliveries = deliveries.order_by("-created_at")
    
    if format_type == "excel":
        return generate_excel_report(deliveries)
    else:
        return generate_pdf_report(deliveries)


def generate_excel_report(deliveries):
    """Genera reporte en formato Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Entregas"
    
    # Encabezados
    headers = [
        "ID Pedido", "Cliente", "Email", "Dirección", "Ciudad",
        "Estado", "Repartidor", "Fecha Programada", "Franja Horaria",
        "Fecha Creación", "Total ($)"
    ]
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="01C9FF", end_color="01C9FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row, delivery in enumerate(deliveries, 2):
        order = delivery.order
        ws.cell(row=row, column=1, value=order.id)
        ws.cell(row=row, column=2, value=order.full_name)
        ws.cell(row=row, column=3, value=order.email)
        ws.cell(row=row, column=4, value=order.address)
        ws.cell(row=row, column=5, value=order.city)
        ws.cell(row=row, column=6, value=delivery.get_status_display())
        ws.cell(row=row, column=7, value=delivery.rider.username if delivery.rider else "Sin asignar")
        ws.cell(row=row, column=8, value=delivery.scheduled_date.strftime("%d/%m/%Y") if delivery.scheduled_date else "")
        ws.cell(row=row, column=9, value=delivery.scheduled_window or "")
        ws.cell(row=row, column=10, value=delivery.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=11, value=float(order.total_amount))
    
    # Ajustar ancho de columnas
    column_widths = [12, 25, 25, 30, 15, 15, 15, 18, 15, 18, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"reporte_entregas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def generate_pdf_report(deliveries):
    """Genera reporte en formato PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type="application/pdf")
    filename = f"reporte_entregas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.toColor("#01C9FF"),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Título
    elements.append(Paragraph("Reporte de Entregas", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elements.append(Spacer(1, 0.3 * inch))
    
    # Preparar datos para la tabla
    data = [["ID", "Cliente", "Estado", "Repartidor", "Fecha Prog.", "Total"]]
    
    for delivery in deliveries:
        order = delivery.order
        row = [
            str(order.id),
            order.full_name[:20] + "..." if len(order.full_name) > 20 else order.full_name,
            delivery.get_status_display(),
            delivery.rider.username[:15] + "..." if delivery.rider and len(delivery.rider.username) > 15 else (delivery.rider.username if delivery.rider else "Sin asignar"),
            delivery.scheduled_date.strftime("%d/%m/%Y") if delivery.scheduled_date else "",
            f"${order.total_amount:.2f}"
        ]
        data.append(row)
    
    # Crear tabla
    table = Table(data, colWidths=[0.8*inch, 2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch])
    
    # Estilo de tabla
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.toColor("#01C9FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Resumen
    elements.append(Spacer(1, 0.3 * inch))
    total_count = deliveries.count()
    delivered_count = deliveries.filter(status="entregada").count()
    failed_count = deliveries.filter(status="fallida").count()
    
    summary_text = f"""
    <b>Resumen:</b><br/>
    Total de entregas: {total_count}<br/>
    Entregadas: {delivered_count}<br/>
    Fallidas: {failed_count}
    """
    elements.append(Paragraph(summary_text, styles["Normal"]))
    
    doc.build(elements)
    return response


@login_required
def failure_statistics(request):
    """Dashboard con estadísticas de causas de fallo en entregas"""
    user = request.user
    if not user.is_superuser and (not hasattr(user, "profile") or user.profile.role != "manager"):
        return redirect("orders:my_orders")
    
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    
    # Filtros opcionales
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    
    # Obtener todas las razones de fallo
    failure_reasons_qs = DeliveryFailureReason.objects.select_related("delivery", "delivery__order", "reported_by")
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            failure_reasons_qs = failure_reasons_qs.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            failure_reasons_qs = failure_reasons_qs.filter(created_at__lte=date_to_obj)
        except ValueError:
            pass
    
    # Estadísticas por razón
    reason_stats = failure_reasons_qs.values("reason").annotate(
        count=Count("id")
    ).order_by("-count")
    
    # Convertir a formato más legible
    reason_stats_list = []
    total_failures = failure_reasons_qs.count()
    
    for stat in reason_stats:
        reason_code = stat["reason"]
        count = stat["count"]
        percentage = (count / total_failures * 100) if total_failures > 0 else 0
        
        reason_stats_list.append({
            "code": reason_code,
            "display": dict(DeliveryFailureReason.FAILURE_REASONS).get(reason_code, reason_code),
            "count": count,
            "percentage": round(percentage, 1)
        })
    
    # Estadísticas por número de intento
    attempt_stats = failure_reasons_qs.values("attempt_number").annotate(
        count=Count("id")
    ).order_by("attempt_number")
    
    # Estadísticas por repartidor
    rider_stats = failure_reasons_qs.values(
        "reported_by__username"
    ).annotate(
        count=Count("id")
    ).order_by("-count")[:10]  # Top 10 repartidores
    
    # Entregas con múltiples fallos
    deliveries_with_multiple_failures = Delivery.objects.annotate(
        failure_count=Count("failure_reasons")
    ).filter(failure_count__gt=1).order_by("-failure_count")[:10]
    
    # Estadísticas por mes (últimos 6 meses) - simplificado
    from django.utils import timezone
    from collections import defaultdict
    six_months_ago = timezone.now() - timedelta(days=180)
    recent_failures = failure_reasons_qs.filter(created_at__gte=six_months_ago)
    
    monthly_dict = defaultdict(int)
    for failure in recent_failures:
        month_key = failure.created_at.strftime("%Y-%m")
        monthly_dict[month_key] += 1
    
    monthly_stats = [{"month": month, "count": count} for month, count in sorted(monthly_dict.items())]
    
    context = {
        "reason_stats": reason_stats_list,
        "total_failures": total_failures,
        "attempt_stats": list(attempt_stats),
        "rider_stats": list(rider_stats),
        "deliveries_with_multiple_failures": deliveries_with_multiple_failures,
        "monthly_stats": list(monthly_stats),
        "date_from": date_from,
        "date_to": date_to,
    }
    
    return render(request, "orders/failure_statistics.html", context)
